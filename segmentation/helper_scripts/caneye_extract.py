
import os
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup

# Set the root directory where all date folders are stored
root_dir = r"./research/dhp_analysis/dhp"  # CHANGE THIS to your root folder path

# List to store all results
records = []

def extract_from_html(html_path):
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            soup = BeautifulSoup(f, 'html.parser')
        
        # Locate all tables and find the one that has "Miller" etc.
        tables = soup.find_all("table")
        target_table = None
        for table in tables:
            if "Miller" in table.text and "LAI2000" in table.text:
                target_table = table
                break

        if not target_table:
            return ["NA"] * 4
        
        rows = target_table.find_all("tr")
        if len(rows) < 2:
            return ["NA"] * 4
        
        # Extract second row, columns 5 to 8 (0-based index)
        data_cells = rows[1].find_all("td")[4:8]
        values = [cell.get_text(strip=True) if cell else "NA" for cell in data_cells]
        return values
    except Exception as e:
        print(f"⚠️ HTML parsing failed: {e}")
        return ["NA"] * 4

# Traverse through folders
for date_folder in os.listdir(root_dir):
    date_path = os.path.join(root_dir, date_folder)
    if not os.path.isdir(date_path):
        continue

    for plot_folder in os.listdir(date_path):
        plot_path = os.path.join(date_path, plot_folder)
        if not os.path.isdir(plot_path):
            continue

        record = {
            "Date": date_folder,
            "Plot": plot_folder,
            "B5": "NA",
            "B6": "NA",
            "B7": "NA",
            "B8": "NA"
        }

        try:
            # Find CE_P180 subfolder
            ce_folder = next((f for f in os.listdir(plot_path) if f.startswith("CE_P180")), None)
            if not ce_folder:
                records.append(record)
                continue

            ce_path = os.path.join(plot_path, ce_folder)

            # Find Excel file
            excel_file = next((f for f in os.listdir(ce_path) if f.startswith("CE_P180") and f.endswith(".xlsx")), None)
            if excel_file:
                excel_path = os.path.join(ce_path, excel_file)

                wb = load_workbook(excel_path, data_only=True)
                if "PAI, ALA" in wb.sheetnames:
                    sheet = wb["PAI, ALA"]
                    record["B5"] = sheet["B5"].value if sheet["B5"].value else "NA"
                    record["B6"] = sheet["B6"].value if sheet["B6"].value else "NA"
                    record["B7"] = sheet["B7"].value if sheet["B7"].value else "NA"
                    record["B8"] = sheet["B8"].value if sheet["B8"].value else "NA"

            # Fallback to HTML if any value still NA
            if any(record[k] == "NA" for k in ["B5", "B6", "B7", "B8"]):
                html_file = next((f for f in os.listdir(ce_path) if f.endswith(".html")), None)
                if html_file:
                    html_path = os.path.join(ce_path, html_file)
                    b_vals = extract_from_html(html_path)
                    record["B5"], record["B6"], record["B7"], record["B8"] = b_vals

        except Exception as e:
            print(f"Error in {date_folder}/{plot_folder}: {e}")

        records.append(record)

# Save results to Excel
df = pd.DataFrame(records)
df.to_excel("CE_P180_Extraction_Summary5.xlsx", index=False)
print("Extraction complete. File saved as 'CE_P180_Extraction_Summary.xlsx'")
